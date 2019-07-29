# Handles Time stamps and time tracking

import openpyxl
import time
import datetime

class TimeTable:
    def __init__(self):
        self.date = datetime.datetime.now()
        self.day = self.date.strftime('%Y/%m/%d')
        self.wb = openpyxl.load_workbook('TimeTable.xlsx')
        self.starttime = "00:00:00"
        self.breaktime = "00:00:00"
        self.endtime = "00:00:00"
        
        self.onBreak = False
        self.breakstart = "00:00:00"
        self.breakend = "00:00:00"
        
        self.lastRow = 0

tm = TimeTable()

def StampStart():
    tm.date = datetime.datetime.now()
    tm.starttime = tm.date.strftime('%H:%M:%S')
    print("Stamp Start: " + tm.starttime)
    
def StampStartBreak():
    tm.date = datetime.datetime.now()
    tm.breakstart = tm.date.strftime('%H:%M:%S')
    tm.onBreak = True
    print("Stamp Start Break: "  + tm.breakstart)
    
def StampEndBreak():
    tm.date = datetime.datetime.now()
    tm.breakend = tm.date.strftime('%H:%M:%S')
    tm.onBreak = False
    print("Stamp End Break: "  + tm.breakend)
    
    diff = datetime.datetime.strptime(tm.breakend, '%H:%M:%S') - datetime.datetime.strptime(tm.breakstart, '%H:%M:%S')
    tm.breaktime = str(diff);
    print("Stamp Break: "  + tm.breaktime)
    

def StampEnd():
    tm.date = datetime.datetime.now()
    tm.endtime = tm.date.strftime('%H:%M:%S')
    print("Stamp End: " + tm.endtime)
    
    LogTime()
    


def LogTime():
    sheet = tm.wb['Sheet1']
    row_ = sheet.max_row
    
    if tm.lastRow != row_ :
        tm.lastRow = row_
    
    while True:
        value = sheet.cell(row= row_, column=1).value
    
        if str(value) == "None":
            break;
        else:
            row_ = row_ + 1
            
        print(value) 
        print(row_)
    
    # Log Date
    sheet.cell(row= row_, column= 1).value = tm.day
    print("Date: " + tm.day + " Logged")
    
    # Log Start Time
    sheet.cell(row= row_, column= 2).value = tm.starttime
    print("Start Time: " + tm.starttime + " Logged")
    
    # Log Break time
    sheet.cell(row= row_, column= 3).value = tm.breaktime
    print("Break Time: " + tm.breaktime + " Logged")
    
    # Log End Time
    sheet.cell(row= row_, column= 4).value = tm.endtime
    print("End Time: " + tm.endtime + " Logged")
    
    tm.wb.save('TimeTable.xlsx')
    
